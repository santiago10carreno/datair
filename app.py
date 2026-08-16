import streamlit as st
import pandas as pd
import requests
import plotly.express as px
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import LineChart, Reference
import concurrent.futures
import math

# ==========================================
# CONFIGURACIÓN Y ESTILOS CORPORATIVOS
# ==========================================
st.set_page_config(page_title="Datair | Inteligencia Ambiental", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');
        html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
        
        /* 1. LA OPCIÓN NUCLEAR: Congelar la caja maestra del Header para siempre */
        .stApp > header, header[data-testid="stHeader"] {
            background-color: transparent !important;
            opacity: 1 !important;
            visibility: visible !important;
            pointer-events: none !important; /* Que el mouse lo traspase */
        }
        
        /* 2. ARRANCAR EL BOTÓN Y PEGARLO A LA PANTALLA (Inmune a la invisibilidad) */
        [data-testid="collapsedControl"], [data-testid="stSidebarCollapsedControl"] {
            position: fixed !important;
            top: 15px !important;
            left: 15px !important;
            z-index: 999999 !important;
            opacity: 1 !important;
            visibility: visible !important;
            display: flex !important;
            pointer-events: auto !important; /* Le devolvemos el poder del clic */
            background-color: #1A1C1E !important;
            border: 1px solid #2D3139 !important;
            border-radius: 8px !important;
            padding: 3px !important;
            color: #E2E8F0 !important;
            transition: none !important;
            transform: none !important;
        }

        /* 3. BORRAR LA BASURA RESTANTE */
        [data-testid="stToolbar"], 
        [data-testid="stDecoration"], 
        .stAppDeployButton, 
        #MainMenu, 
        footer {
            display: none !important;
            visibility: hidden !important;
        }

        /* 4. DISEÑO DE LAS MÉTRICAS */
        div[data-testid="metric-container"] {
            background-color: #1A1C1E;
            border: 1px solid #2D3139;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
            display: flex !important;
            flex-direction: column !important;
            align-items: flex-start !important;
            justify-content: center !important;
            width: 100% !important;
        }
        
        [data-testid="stMetricLabel"], [data-testid="stMetricValue"] {
            text-align: left !important;
            width: 100% !important;
        }
        [data-testid="stMetricValue"] > div {
            white-space: normal !important;
            word-wrap: break-word !important;
            font-size: 1.6rem !important;
            line-height: 1.2 !important;
            text-align: left !important;
        }

        h1 { color: #4A90E2 !important; font-weight: 700 !important; letter-spacing: -0.5px; }
        h2, h3 { font-weight: 600 !important; letter-spacing: -0.3px; color: #E2E8F0 !important; }
        .stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
            font-size: 1.05rem; font-weight: 600;
        }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 1. BASE DE DATOS GEOGRÁFICA Y HARDWARE
# ==========================================
DICCIONARIO_ZONAS = {
    "Región de Antofagasta": {"Antofagasta": {"Antofagasta (Centro)": (-23.6500, -70.4000)}, "Calama": {"Calama": (-22.4500, -68.9300)}},
    "Región de Valparaíso": {"Valparaíso": {"Valparaíso": (-33.0500, -71.6200)}, "Viña del Mar": {"Viña del Mar": (-33.0200, -71.5500)}, "Quintero": {"Quintero (Centro)": (-32.7800, -71.5300), "Loncura": (-32.7900, -71.5200)}, "Puchuncaví": {"Puchuncaví": (-32.7300, -71.4100), "Las Ventanas": (-32.7500, -71.4800)}},
    "Región Metropolitana": {"Santiago Centro": {"Parque O'Higgins": (-33.4641, -70.6607)}, "Independencia": {"Independencia": (-33.4150, -70.6528)}, "Pudahuel": {"Pudahuel": (-33.4326, -70.7818)}, "Quilicura": {"Quilicura": (-33.3663, -70.7351)}, "Las Condes": {"Las Condes": (-33.3769, -70.5239)}, "Cerrillos": {"Cerrillos": (-33.4939, -70.7161)}, "El Bosque": {"El Bosque": (-33.5350, -70.6766)}, "Cerro Navia": {"Cerro Navia": (-33.4334, -70.7348)}, "Puente Alto": {"Puente Alto": (-33.6163, -70.5831)}, "Talagante": {"Talagante": (-33.6669, -70.9275)}},
    "Región de O'Higgins": {"Rancagua": {"Rancagua (Centro)": (-34.1708, -70.7441), "Rancagua 2 (Norte)": (-34.1522, -70.7305)}, "Rengo": {"Rengo (Centro)": (-34.4091, -70.8591)}, "San Fernando": {"San Fernando": (-34.5847, -70.9880)}, "Machalí": {"Machalí": (-34.1816, -70.6558)}},
    "Región del Maule": {"Talca": {"Talca (Centro)": (-35.4300, -71.6700), "La Florida": (-35.4500, -71.6800)}, "Curicó": {"Curicó": (-34.9800, -71.2300)}},
    "Región del Biobío": {"Concepción": {"Concepción": (-36.8300, -73.0500)}, "Talcahuano": {"Talcahuano": (-36.7200, -73.1200)}, "Coronel": {"Coronel Norte": (-37.0100, -73.1400), "Coronel Sur": (-37.0400, -73.1500)}, "Los Ángeles": {"Los Ángeles": (-37.4700, -72.3500)}},
    "Región de La Araucanía": {"Temuco": {"Temuco (Centro)": (-38.7400, -72.5900)}, "Padre Las Casas": {"Padre Las Casas": (-38.7700, -72.5900)}},
    "Región de Los Ríos": {"Valdivia": {"Valdivia": (-39.8100, -73.2400)}},
    "Región de Los Lagos": {"Osorno": {"Osorno": (-40.5700, -73.1300)}, "Puerto Montt": {"Puerto Montt": (-41.4700, -72.9300)}},
    "Región de Aysén": {"Coyhaique": {"Coyhaique 1": (-45.5700, -72.0700), "Coyhaique 2": (-45.5800, -72.0600)}},
    "Región de Magallanes": {"Punta Arenas": {"Punta Arenas": (-53.1500, -70.9000)}}
}

MAPA_SENSORES = {
    "Quintero (Centro)": ["MP2.5", "MP10", "SO2", "NO2", "O3", "CO"],
    "Loncura": ["SO2", "NO2"],
    "Las Ventanas": ["MP10", "SO2", "NO2"],
    "Parque O'Higgins": ["MP2.5", "MP10", "SO2", "CO", "NO2", "O3"],
    "Pudahuel": ["MP2.5", "MP10", "O3"],
    "Quilicura": ["MP2.5", "MP10", "NO2"],
    "Independencia": ["MP2.5", "MP10", "CO"],
    "Rancagua (Centro)": ["MP2.5", "MP10", "CO"],
    "Coronel Norte": ["MP2.5", "MP10", "SO2", "NO2"],
    "Coronel Sur": ["MP10", "SO2"],
    "Antofagasta (Centro)": ["MP2.5", "MP10", "SO2"],
    "Calama": ["MP10", "SO2"]
}

def obtener_sensores_certificados(estacion):
    return MAPA_SENSORES.get(estacion, ["MP2.5", "MP10"])

configuracion = {
    "MP2.5": {"api": "pm2_5", "limite": 50.0},
    "MP10": {"api": "pm10", "limite": 150.0},
    "SO2": {"api": "sulphur_dioxide", "limite": 250.0},
    "CO": {"api": "carbon_monoxide", "limite": 10000.0}, 
    "NO2": {"api": "nitrogen_dioxide", "limite": 400.0}, 
    "O3": {"api": "ozone", "limite": 120.0}
}

# ==========================================
# 2. MOTOR ICAP CHILE
# ==========================================
def evaluar_icap(valor, contaminante):
    if contaminante == "MP2.5":
        if valor <= 50: return "Bueno", "#00E400", 8
        elif valor <= 79: return "Regular", "#FFFF00", 10
        elif valor <= 109: return "Alerta", "#FF7E00", 13
        elif valor <= 169: return "Preemergencia", "#FF0000", 16
        else: return "Emergencia", "#8F3F97", 22
        
    elif contaminante == "MP10":
        if valor <= 150: return "Bueno", "#00E400", 8
        elif valor <= 194: return "Regular", "#FFFF00", 10
        elif valor <= 239: return "Alerta", "#FF7E00", 13
        elif valor <= 329: return "Preemergencia", "#FF0000", 16
        else: return "Emergencia", "#8F3F97", 22
        
    else:
        limite = configuracion[contaminante]["limite"]
        pct = valor / limite
        if pct <= 0.6: return "Bueno", "#00E400", 8
        elif pct <= 0.9: return "Regular", "#FFFF00", 10
        elif pct <= 1.2: return "Alerta", "#FF7E00", 13
        elif pct <= 1.5: return "Preemergencia", "#FF0000", 16
        else: return "Emergencia", "#8F3F97", 22

PALETA_ICAP = {
    "Bueno": "#00E400", 
    "Regular": "#FFFF00", 
    "Alerta": "#FF7E00", 
    "Preemergencia": "#FF0000", 
    "Emergencia": "#8F3F97"
}

# ==========================================
# 3. BARRA LATERAL 
# ==========================================
st.sidebar.title("Configuracion Global")
contaminante_elegido = st.sidebar.selectbox("Selecciona el Contaminante Principal", list(configuracion.keys()))

var_api = configuracion[contaminante_elegido]["api"]
limite_actual = configuracion[contaminante_elegido]["limite"]

# ==========================================
# 4. EXTRACCIÓN ASÍNCRONA
# ==========================================
def obtener_datos_estacion_individual(args):
    lat, lon, variable, region, comuna, sector = args
    try:
        url = f"https://air-quality-api.open-meteo.com/v1/air-quality?latitude={lat}&longitude={lon}&hourly={variable}&timezone=America%2FSantiago&past_days=7&forecast_days=3"
        respuesta = requests.get(url, timeout=5)
        datos = respuesta.json()
        fechas = pd.to_datetime(datos['hourly']['time']).tz_localize(None)
        serie = pd.Series(datos['hourly'][variable], index=fechas)
        return (region, comuna, sector, serie)
    except:
        return (region, comuna, sector, pd.Series(dtype=float))

@st.cache_data(ttl=3600)
def descargar_todos_los_datos(contaminante_nombre, variable_api):
    lista_tareas = []
    estaciones_con_hardware = 0
    for region, comunas in DICCIONARIO_ZONAS.items():
        for comuna, sectores in comunas.items():
            for sector, coords in sectores.items():
                if contaminante_nombre in obtener_sensores_certificados(sector):
                    estaciones_con_hardware += 1
                    lista_tareas.append((coords[0], coords[1], variable_api, region, comuna, sector))
    
    resultados_completos = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        for region, comuna, sector, serie in executor.map(obtener_datos_estacion_individual, lista_tareas):
            if not serie.empty:
                if region not in resultados_completos: resultados_completos[region] = {}
                if comuna not in resultados_completos[region]: resultados_completos[region][comuna] = {}
                resultados_completos[region][comuna][sector] = serie
    return resultados_completos, estaciones_con_hardware

@st.cache_data(ttl=3600)
def obtener_datos_multivariable(lat, lon, variables_api_lista):
    variables_str = ",".join(variables_api_lista)
    try:
        url = f"https://air-quality-api.open-meteo.com/v1/air-quality?latitude={lat}&longitude={lon}&hourly={variables_str}&timezone=America%2FSantiago&past_days=7&forecast_days=3"
        res = requests.get(url, timeout=5)
        datos = res.json()
        fechas = pd.to_datetime(datos['hourly']['time']).tz_localize(None)
        df_multi = pd.DataFrame(index=fechas)
        for var in variables_api_lista: df_multi[var] = datos['hourly'][var]
        return df_multi
    except:
        return pd.DataFrame()

@st.cache_data(ttl=3600)
def obtener_meteorologia(lat, lon):
    try:
        url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&hourly=temperature_2m,wind_speed_10m,wind_direction_10m&timezone=America%2FSantiago&past_days=7&forecast_days=3"
        res = requests.get(url, timeout=5)
        datos = res.json()
        fechas = pd.to_datetime(datos['hourly']['time']).tz_localize(None)
        df_met = pd.DataFrame({
            'Temperatura (°C)': datos['hourly']['temperature_2m'],
            'Velocidad Viento (km/h)': datos['hourly']['wind_speed_10m'],
            'Direccion Viento (°)': datos['hourly']['wind_direction_10m']
        }, index=fechas)
        return df_met.reset_index().rename(columns={'index': 'Fecha y Hora'})
    except:
        return pd.DataFrame()

def obtener_viento_batch(df):
    if df.empty: return df
    lats = ",".join(df['Latitud'].astype(str))
    lons = ",".join(df['Longitud'].astype(str))
    url = f"https://api.open-meteo.com/v1/forecast?latitude={lats}&longitude={lons}&current=wind_speed_10m,wind_direction_10m&timezone=America%2FSantiago"
    try:
        res = requests.get(url, timeout=5)
        data = res.json()
        vels, dirs = [], []
        if isinstance(data, list):
            for d in data:
                vels.append(d.get('current', {}).get('wind_speed_10m', 0))
                dirs.append(d.get('current', {}).get('wind_direction_10m', 0))
        else:
            vels.append(data.get('current', {}).get('wind_speed_10m', 0))
            dirs.append(data.get('current', {}).get('wind_direction_10m', 0))
        df['WindSpd'] = vels
        df['WindDir'] = dirs
    except:
        df['WindSpd'] = 0
        df['WindDir'] = 0
    return df

datos_totales, total_hardware_valido = descargar_todos_los_datos(contaminante_elegido, var_api)
ahora = pd.Timestamp.now(tz='America/Santiago').tz_localize(None)

# ==========================================
# 5. PREPARACIÓN DE DATOS CON LÓGICA ICAP
# ==========================================
datos_mapa = []
for region, comunas in datos_totales.items():
    for comuna, sectores in comunas.items():
        for sector, serie in sectores.items():
            if not serie.empty:
                try:
                    idx_actual = serie.index.get_indexer([ahora], method='nearest')[0]
                    valor_actual = serie.iloc[idx_actual]
                except:
                    valor_actual = serie.iloc[-1]
                
                estado_icap, color_icap, tamanio_icap = evaluar_icap(valor_actual, contaminante_elegido)
                
                datos_mapa.append({
                    "Region": region, "Comuna": comuna, "Estacion": sector,
                    "Latitud": DICCIONARIO_ZONAS[region][comuna][sector][0], "Longitud": DICCIONARIO_ZONAS[region][comuna][sector][1],
                    "Concentracion": round(valor_actual, 1), 
                    "Estado": estado_icap, 
                    "Color": color_icap, 
                    "Tamaño": tamanio_icap
                })
df_mapa = pd.DataFrame(datos_mapa)

# ==========================================
# 6. UI PRINCIPAL Y KPIS
# ==========================================
st.title("Datair | Inteligencia Ambiental")
st.markdown(f"**Monitoreo Nacional de {contaminante_elegido}** | Limite Normativo 24h: `{limite_actual} µg/m³`")

promedio_nacional = df_mapa["Concentracion"].mean() if not df_mapa.empty else 0
estaciones_criticas = len(df_mapa[df_mapa["Estado"].isin(["Alerta", "Preemergencia", "Emergencia"])]) if not df_mapa.empty else 0

if estaciones_criticas > (len(df_mapa) * 0.3): estado_pais = "Emergencia Nacional"
elif estaciones_criticas > 0: estado_pais = "Zonas en Riesgo"
else: estado_pais = "Condiciones Óptimas"

col1, col2, col3 = st.columns(3)
with col1: st.metric(label="Promedio Pais Actual", value=f"{promedio_nacional:.1f} µg/m³")
with col2: st.metric(label="Estado General", value=estado_pais)
with col3: st.metric(label="Sectores en Riesgo (ICAP)", value=f"{estaciones_criticas} de {total_hardware_valido}", help="Ve a la pestaña 'Centro de Alertas' para ver el detalle gráfico de las zonas en riesgo.")

if estaciones_criticas > 0:
    with st.expander("🚨 Ver detalle de los Sectores en Riesgo actuales (Aviso: Ve a la pestaña 'Centro de Alertas' para gráficos profundos)"):
        df_riesgo = df_mapa[df_mapa["Estado"].isin(["Alerta", "Preemergencia", "Emergencia"])].copy()
        df_riesgo = df_riesgo.sort_values(by="Concentracion", ascending=False).reset_index(drop=True)
        st.dataframe(df_riesgo[["Region", "Comuna", "Estacion", "Concentracion", "Estado"]], use_container_width=True, hide_index=True)

st.divider()

st.markdown("""
<div style="display: flex; justify-content: center; gap: 15px; margin-bottom: 20px; font-size: 0.9rem;">
    <div><span style="color:#00E400;">🟢</span> Bueno</div>
    <div><span style="color:#FFFF00;">🟡</span> Regular</div>
    <div><span style="color:#FF7E00;">🟠</span> Alerta</div>
    <div><span style="color:#FF0000;">🔴</span> Preemergencia</div>
    <div><span style="color:#8F3F97;">🟣</span> Emergencia</div>
</div>
""", unsafe_allow_html=True)

# ==========================================
# SISTEMA DE PESTAÑAS (TABS)
# ==========================================
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "Monitoreo y Analisis", "Proyeccion", "Benchmarking", "Multivariable", "Meteorologia", "Centro de Alertas"
])

# ------------------------------------------
# TAB 1: MONITOREO Y DISPERSIÓN
# ------------------------------------------
with tab1:
    st.subheader("Red de Monitoreo (Semáforo ICAP)")
    if not df_mapa.empty:
        fig_mapa = px.scatter_mapbox(
            df_mapa, lat="Latitud", lon="Longitud", hover_name="Estacion", 
            hover_data={"Latitud": False, "Longitud": False, "Region": True, "Comuna": True, "Concentracion": True, "Estado": True, "Tamaño": False, "Color": False},
            color="Estado", color_discrete_map=PALETA_ICAP,
            size="Tamaño", zoom=4, center={"lat": -35.0, "lon": -71.0}, height=500,
            category_orders={"Estado": ["Bueno", "Regular", "Alerta", "Preemergencia", "Emergencia"]}
        )
        fig_mapa.update_layout(mapbox_style="carto-darkmatter", margin={"r":0,"t":0,"l":0,"b":0})
        st.plotly_chart(fig_mapa, use_container_width=True)
    else:
        st.warning(f"No hay estaciones certificadas con sensores de {contaminante_elegido} respondiendo actualmente.")
    
    st.divider()
    
    st.subheader("Modelo de Dispersion Espacial Organico")
    if not df_mapa.empty:
        df_vectores = obtener_viento_batch(df_mapa.copy())
        
        puntos_pluma = []
        for _, row in df_vectores.iterrows():
            lat = row['Latitud']
            lon = row['Longitud']
            c = row['Concentracion']
            spd = row.get('WindSpd', 0)
            dir_viento = row.get('WindDir', 0)
            
            puntos_pluma.append(row.to_dict())
            
            if spd > 1:
                angulo_viaje = (dir_viento + 180) % 360
                angulo_rad = math.radians(90 - angulo_viaje)
                pasos = 8
                dist_max = spd * 0.015 
                for i in range(1, pasos + 1):
                    frac = i / pasos
                    dist = dist_max * frac
                    d_lat = dist * math.sin(angulo_rad)
                    d_lon = dist * math.cos(angulo_rad) / math.cos(math.radians(lat))
                    
                    c_fantasma = c * (1 - frac)**1.5 
                    if c_fantasma > (limite_actual * 0.05):
                        nuevo_punto = row.to_dict()
                        nuevo_punto['Latitud'] = lat + d_lat
                        nuevo_punto['Longitud'] = lon + d_lon
                        nuevo_punto['Concentracion'] = c_fantasma
                        nuevo_punto['Estacion'] = f"Viento desde {row['Estacion']}"
                        puntos_pluma.append(nuevo_punto)
        
        df_pluma = pd.DataFrame(puntos_pluma)
        regiones_disp = list(df_vectores['Region'].unique())
        reg_mapa = st.selectbox("Enfocar cámara en la Región:", regiones_disp, index=0 if "Región Metropolitana" not in regiones_disp else regiones_disp.index("Región Metropolitana"), key="heatmap_region")
        
        df_region_mapa = df_vectores[df_vectores['Region'] == reg_mapa]
        if not df_region_mapa.empty:
            lat_centro = df_region_mapa['Latitud'].mean()
            lon_centro = df_region_mapa['Longitud'].mean()
        else:
            lat_centro, lon_centro = -35.0, -71.0

        fig_heat = px.density_mapbox(
            df_pluma, lat="Latitud", lon="Longitud", z="Concentracion",
            radius=60, center={"lat": lat_centro, "lon": lon_centro}, 
            zoom=9, mapbox_style="carto-darkmatter", color_continuous_scale="Inferno", 
            opacity=0.6, hover_name="Estacion"
        )
        fig_heat.update_layout(margin={"r":0,"t":40,"l":0,"b":0})
        st.plotly_chart(fig_heat, use_container_width=True)

    st.divider()
    
    st.subheader("Filtros de Analisis Local Historico")
    regiones_disponibles = list(datos_totales.keys())
    if regiones_disponibles:
        col_filtro1, col_filtro2 = st.columns(2)
        with col_filtro1: region_elegida = st.selectbox("Selecciona la Region", regiones_disponibles, key="reg_hist")
        with col_filtro2: comuna_elegida = st.selectbox("Selecciona la Comuna", list(datos_totales[region_elegida].keys()), key="com_hist")

        datos_sectores_comuna = {}
        lista_promedios_comunas = {}

        for comuna, sectores in datos_totales[region_elegida].items():
            series_comuna = []
            for sector, serie in sectores.items():
                series_comuna.append(serie)
                if comuna == comuna_elegida:
                    datos_sectores_comuna[sector] = serie
            if series_comuna:
                lista_promedios_comunas[comuna] = pd.concat(series_comuna, axis=1).mean(axis=1)

        df_region_completo = pd.DataFrame(lista_promedios_comunas).reset_index().rename(columns={'index': 'Fecha y Hora'})
        df_comuna_completo = pd.DataFrame(datos_sectores_comuna).reset_index().rename(columns={'index': 'Fecha y Hora'})
        df_region_historico = df_region_completo[df_region_completo['Fecha y Hora'] <= ahora]
        df_comuna_historico = df_comuna_completo[df_comuna_completo['Fecha y Hora'] <= ahora]

        st.subheader(f"Analisis Historico Regional: {region_elegida}")
        fig_reg_hist = px.line(df_region_historico, x='Fecha y Hora', y=df_region_historico.columns[1:], labels={'value': 'Concentracion (µg/m³)', 'variable': 'Comuna'})
        fig_reg_hist.add_hline(y=limite_actual, line_dash="dot", line_color="red", annotation_text="Limite Legal")
        st.plotly_chart(fig_reg_hist, use_container_width=True)

        st.subheader(f"Analisis Historico Comunal: {comuna_elegida}")
        fig_com_hist = px.line(df_comuna_historico, x='Fecha y Hora', y=df_comuna_historico.columns[1:], labels={'value': 'Concentracion (µg/m³)', 'variable': 'Estacion'})
        fig_com_hist.add_hline(y=limite_actual, line_dash="dot", line_color="red", annotation_text="Limite Legal")
        st.plotly_chart(fig_com_hist, use_container_width=True)
        
        st.divider()
        
        st.subheader("Generacion de Reportes de Cumplimiento")
        st.write("Descarga informes gerenciales auditables.")

        def generar_excel_universal(df_datos, contaminante, limite, nombre_zona, tipo_zona):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_datos.to_excel(writer, sheet_name='Base_Datos', index=False, startrow=3)
                ws_resumen = writer.book.create_sheet('Dashboard_Ejecutivo', 0)
                ws_datos = writer.sheets['Base_Datos']
                
                ws_datos['A1'] = f"REGISTRO CONTINUO - {contaminante} ({nombre_zona})"
                ws_datos['A1'].font = Font(size=12, bold=True)
                ws_datos['A2'] = f"Limite Normativo: {limite} µg/m³"
                ws_datos['A2'].font = Font(italic=True, color="595959")
                
                header_fill, header_font = PatternFill(start_color="2F75B5", fill_type="solid"), Font(bold=True, color="FFFFFF")
                columnas_datos = list(df_datos.columns)[1:] 
                
                for col_idx, _ in enumerate(df_datos.columns, start=1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws_datos[f'{col_letter}4'].fill = header_fill
                    ws_datos[f'{col_letter}4'].font = header_font
                    ws_datos.column_dimensions[col_letter].width = 22
                    
                red_fill, red_font = PatternFill(start_color="FFC7CE", fill_type="solid"), Font(color="9C0006", bold=True)
                green_fill, green_font = PatternFill(start_color="C6EFCE", fill_type="solid"), Font(color="006100")
                
                rule_over = CellIsRule(operator='greaterThan', formula=[str(limite)], stopIfTrue=True, fill=red_fill, font=red_font)
                rule_under = CellIsRule(operator='lessThanOrEqual', formula=[str(limite)], stopIfTrue=True, fill=green_fill, font=green_font)
                
                ultima_letra = openpyxl.utils.get_column_letter(len(df_datos.columns))
                ws_datos.conditional_formatting.add(f'B5:{ultima_letra}{len(df_datos)+4}', rule_over)
                ws_formatting = ws_datos.conditional_formatting
                ws_formatting.add(f'B5:{ultima_letra}{len(df_datos)+4}', rule_under)
                ws_datos.auto_filter.ref = f"A4:{ultima_letra}{len(df_datos)+4}"
                ws_datos.freeze_panes = 'A5'
                
                for row in range(1, 30):
                    for col in range(1, 15):
                        ws_resumen.cell(row=row, column=col).fill = PatternFill(start_color="F2F2F2", fill_type="solid")
                
                ws_resumen['B2'] = f"REPORTE DE AUDITORIA - {tipo_zona.upper()}"
                ws_resumen['B2'].font = Font(size=18, bold=True, color="FFFFFF")
                ws_resumen['B2'].fill = PatternFill(start_color="1F4E78", fill_type="solid")
                ws_resumen.merge_cells('B2:H3')
                ws_resumen['B2'].alignment = Alignment(horizontal="center", vertical="center")
                
                ws_resumen['B4'] = "Tipo de Informe:"
                ws_resumen['C4'] = "Auditoria Oficial"
                ws_resumen['B5'], ws_resumen['C5'] = f"{tipo_zona} Evaluada:", nombre_zona
                ws_resumen['B6'], ws_resumen['C6'] = "Parametro Evaluado:", contaminante
                ws_resumen['B7'], ws_resumen['C7'] = "Limite Legal:", f"{limite} µg/m³"
                ws_resumen['C7'].font = Font(bold=True, color="C00000")
                
                kpi_headers = ['Sub-zona', 'Promedio', 'Peak Maximo', 'Horas Infraccion', 'Estado']
                for i, header in enumerate(kpi_headers, start=2):
                    cell = ws_resumen.cell(row=10, column=i)
                    cell.value, cell.font, cell.fill, cell.alignment = header, header_font, header_fill, Alignment(horizontal="center")
                    ws_resumen.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20
                    
                for idx, subzona in enumerate(columnas_datos):
                    row = 11 + idx
                    horas_infraccion = (df_datos[subzona] > limite).sum()
                    ws_resumen.cell(row=row, column=2).value = subzona
                    ws_resumen.cell(row=row, column=3).value = round(df_datos[subzona].mean(), 2)
                    ws_resumen.cell(row=row, column=4).value = round(df_datos[subzona].max(), 2)
                    ws_resumen.cell(row=row, column=5).value = horas_infraccion
                    estado_cell = ws_resumen.cell(row=row, column=6)
                    if horas_infraccion > 0:
                        estado_cell.value, estado_cell.font, estado_cell.fill = "CRITICO", red_font, red_fill
                    else:
                        estado_cell.value, estado_cell.font, estado_cell.fill = "CUMPLE", green_font, green_fill
                        
                    for col in range(2, 7):
                        ws_resumen.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                        ws_resumen.cell(row=row, column=col).border = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'), top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))
                        
                chart = LineChart()
                chart.title = f"Monitoreo Continuo - {contaminante}"
                chart.style, chart.width, chart.height = 13, 25, 13
                data = Reference(ws_datos, min_col=2, min_row=4, max_col=len(df_datos.columns), max_row=len(df_datos)+4)
                cats = Reference(ws_datos, min_col=1, min_row=5, max_row=len(df_datos)+4)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                ws_resumen.add_chart(chart, "B16")
                
            return output.getvalue()

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            excel_region = generar_excel_universal(df_region_historico, contaminante_elegido, limite_actual, region_elegida, "Region")
            st.download_button(
                label=f"Descargar Auditoria Regional ({region_elegida})",
                data=excel_region,
                file_name=f"Auditoria_{region_elegida}_{contaminante_elegido}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col_btn2:
            excel_comuna = generar_excel_universal(df_comuna_historico, contaminante_elegido, limite_actual, comuna_elegida, "Comuna")
            st.download_button(
                label=f"Descargar Auditoria Comunal ({comuna_elegida})",
                data=excel_comuna,
                file_name=f"Auditoria_{comuna_elegida}_{contaminante_elegido}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

with tab2:
    st.subheader("Filtros de Proyeccion Predictiva")
    if datos_totales:
        col_proy1, col_proy2 = st.columns(2)
        with col_proy1: reg_proy = st.selectbox("Selecciona la Region", list(datos_totales.keys()), key="reg_proy")
        with col_proy2: com_proy = st.selectbox("Selecciona la Comuna", list(datos_totales[reg_proy].keys()), key="com_proy")

        datos_sectores_proy = {}
        lista_promedios_proy = {}
        for comuna, sectores in datos_totales[reg_proy].items():
            series_comuna = []
            for sector, serie in sectores.items():
                series_comuna.append(serie)
                if comuna == com_proy: datos_sectores_proy[sector] = serie
            if series_comuna: lista_promedios_proy[comuna] = pd.concat(series_comuna, axis=1).mean(axis=1)

        df_region_proy = pd.DataFrame(lista_promedios_proy).reset_index().rename(columns={'index': 'Fecha y Hora'})
        df_comuna_proy = pd.DataFrame(datos_sectores_proy).reset_index().rename(columns={'index': 'Fecha y Hora'})

        st.subheader(f"Modelo Predictivo Regional: {reg_proy}")
        fig_region_pred = px.line(df_region_proy, x='Fecha y Hora', y=df_region_proy.columns[1:], labels={'value': 'Concentracion (µg/m³)', 'variable': 'Comuna'})
        fig_region_pred.add_hline(y=limite_actual, line_dash="dot", line_color="red", annotation_text="Limite Legal")
        fig_region_pred.add_vline(x=ahora, line_width=2, line_dash="dash", line_color="white", annotation_text="AHORA")
        st.plotly_chart(fig_region_pred, use_container_width=True)

        st.subheader(f"Modelo Predictivo Comunal: {com_proy}")
        fig_comuna_pred = px.line(df_comuna_proy, x='Fecha y Hora', y=df_comuna_proy.columns[1:], labels={'value': 'Concentracion (µg/m³)', 'variable': 'Estacion'})
        fig_comuna_pred.add_hline(y=limite_actual, line_dash="dot", line_color="red", annotation_text="Limite Legal")
        fig_comuna_pred.add_vline(x=ahora, line_width=2, line_dash="dash", line_color="white", annotation_text="AHORA")
        st.plotly_chart(fig_comuna_pred, use_container_width=True)

with tab3:
    st.subheader("Benchmarking Corporativo (Comparador Cruzado)")
    if len(datos_totales) > 0:
        col_vs1, col_vs2 = st.columns(2)
        with col_vs1:
            reg_a = st.selectbox("Region A", list(datos_totales.keys()), key="reg_a")
            com_a = st.selectbox("Comuna A", list(datos_totales[reg_a].keys()), key="com_a")
        with col_vs2:
            reg_b = st.selectbox("Region B", list(datos_totales.keys()), index=min(1, len(datos_totales)-1), key="reg_b")
            com_b = st.selectbox("Comuna B", list(datos_totales[reg_b].keys()), key="com_b")

        def obtener_promedio_comuna(region, comuna):
            series = []
            if region in datos_totales and comuna in datos_totales[region]:
                for sector, serie in datos_totales[region][comuna].items():
                    if not serie.empty: series.append(serie)
            if series: return pd.concat(series, axis=1).mean(axis=1)
            return pd.Series(dtype=float)

        promedio_a = obtener_promedio_comuna(reg_a, com_a)
        promedio_b = obtener_promedio_comuna(reg_b, com_b)

        if not promedio_a.empty and not promedio_b.empty:
            df_vs = pd.DataFrame({com_a: promedio_a, com_b: promedio_b}).reset_index().rename(columns={'index': 'Fecha y Hora'})
            fig_vs = px.line(df_vs, x='Fecha y Hora', y=[com_a, com_b], labels={'value': 'Concentracion (µg/m³)', 'variable': 'Comuna Analizada'})
            fig_vs.add_hline(y=limite_actual, line_dash="dot", line_color="red", annotation_text="Limite Legal")
            fig_vs.add_vline(x=ahora, line_width=2, line_dash="dash", line_color="white", annotation_text="AHORA")
            st.plotly_chart(fig_vs, use_container_width=True)

with tab4:
    st.subheader("Perfil de Estacion (Analisis Multivariable)")
    col_p1, col_p2, col_p3 = st.columns(3)
    with col_p1: reg_p = st.selectbox("Region", list(DICCIONARIO_ZONAS.keys()), index=3, key="reg_p_multi")
    with col_p2: com_p = st.selectbox("Comuna", list(DICCIONARIO_ZONAS[reg_p].keys()), key="com_p_multi")
    with col_p3: est_p = st.selectbox("Estacion Especifica", list(DICCIONARIO_ZONAS[reg_p][com_p].keys()), key="est_p_multi")

    sensores_certificados = obtener_sensores_certificados(est_p)
    contaminantes_seleccionados = st.multiselect(
        "Selecciona los contaminantes a comparar:", sensores_certificados,  
        default=[s for s in ["MP2.5", "MP10"] if s in sensores_certificados]
    )

    modo_vista = st.radio("Modo de Visualizacion:", ["Grafico Unificado (Porcentaje del Limite)", "Graficos Separados (Absolutos)"], horizontal=True)

    if contaminantes_seleccionados:
        lat_p, lon_p = DICCIONARIO_ZONAS[reg_p][com_p][est_p]
        vars_api = [configuracion[c]["api"] for c in contaminantes_seleccionados]
        df_multi = obtener_datos_multivariable(lat_p, lon_p, vars_api)
        
        if not df_multi.empty:
            df_multi = df_multi.rename(columns={configuracion[c]["api"]: c for c in contaminantes_seleccionados}).reset_index().rename(columns={'index': 'Fecha y Hora'})
            if "Separados" in modo_vista:
                df_melt = df_multi.melt(id_vars=['Fecha y Hora'], value_vars=contaminantes_seleccionados, var_name='Contaminante', value_name='Concentracion')
                fig_multi = px.line(df_melt, x='Fecha y Hora', y='Concentracion', facet_row='Contaminante', height=250 * len(contaminantes_seleccionados))
                fig_multi.update_yaxes(matches=None)
                fig_multi.add_vline(x=ahora, line_width=2, line_dash="dash", line_color="white", annotation_text="AHORA")
                st.plotly_chart(fig_multi, use_container_width=True)
            else:
                df_norm = df_multi.copy()
                for c in contaminantes_seleccionados: df_norm[c] = (df_norm[c] / configuracion[c]["limite"]) * 100
                df_melt_norm = df_norm.melt(id_vars=['Fecha y Hora'], value_vars=contaminantes_seleccionados, var_name='Contaminante', value_name='Porcentaje del Limite (%)')
                fig_norm = px.line(df_melt_norm, x='Fecha y Hora', y='Porcentaje del Limite (%)', color='Contaminante')
                fig_norm.add_hline(y=100, line_dash="dot", line_color="red", annotation_text="LIMITE LEGAL (100%)")
                fig_norm.add_vline(x=ahora, line_width=2, line_dash="dash", line_color="white", annotation_text="AHORA")
                st.plotly_chart(fig_norm, use_container_width=True)

with tab5:
    st.subheader("Contexto Meteorologico y Direccion del Viento")
    col_m1, col_m2, col_m3 = st.columns(3)
    with col_m1: reg_m = st.selectbox("Region", list(DICCIONARIO_ZONAS.keys()), index=3, key="reg_m")
    with col_m2: com_m = st.selectbox("Comuna", list(DICCIONARIO_ZONAS[reg_m].keys()), key="com_m")
    with col_m3: est_m = st.selectbox("Estacion Especifica", list(DICCIONARIO_ZONAS[reg_m][com_m].keys()), key="est_m")

    lat_m, lon_m = DICCIONARIO_ZONAS[reg_m][com_m][est_m]
    df_met = obtener_meteorologia(lat_m, lon_m)
    
    if not df_met.empty:
        try:
            idx_actual_met = df_met['Fecha y Hora'].get_indexer([ahora], method='nearest')[0]
            temp_actual = df_met.iloc[idx_actual_met]['Temperatura (°C)']
            viento_actual = df_met.iloc[idx_actual_met]['Velocidad Viento (km/h)']
            dir_grados = df_met.iloc[idx_actual_met]['Direccion Viento (°)']
            def grados_a_cardinal_local(d):
                dirs = ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE", "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]
                ix = int((d + 11.25)/22.5)
                return dirs[ix % 16]
            dir_cardinal = grados_a_cardinal_local(dir_grados)
        except:
            temp_actual, viento_actual, dir_grados, dir_cardinal = 0, 0, 0, "N/A"

        st.markdown(f"**Condiciones Actuales en {est_m}**")
        col_k1, col_k2, col_k3, col_k4 = st.columns(4)
        with col_k1: st.metric("🌡️ Temperatura", f"{temp_actual:.1f} °C")
        with col_k2: st.metric("🌬️ Velocidad Viento", f"{viento_actual:.1f} km/h")
        with col_k3: st.metric("🧭 Direccion Viento", f"{dir_cardinal} ({dir_grados}°)")
        with col_k4:
            val_c_str = "Sin datos"
            if reg_m in datos_totales and com_m in datos_totales[reg_m] and est_m in datos_totales[reg_m][com_m]:
                serie_cont = datos_totales[reg_m][com_m][est_m]
                if not serie_cont.empty:
                    try:
                        idx_c = serie_cont.index.get_indexer([ahora], method='nearest')[0]
                        val_c = serie_cont.iloc[idx_c]
                        val_c_str = f"{val_c:.1f} µg/m³"
                    except: pass
            st.metric(f"☁️ {contaminante_elegido}", val_c_str)

        st.divider()
        fig_temp = px.line(df_met, x='Fecha y Hora', y='Temperatura (°C)')
        fig_temp.add_vline(x=ahora, line_width=2, line_dash="dash", line_color="white", annotation_text="AHORA")
        fig_temp.update_traces(line_color="#FFA500")
        st.plotly_chart(fig_temp, use_container_width=True)

        fig_viento = px.line(df_met, x='Fecha y Hora', y='Velocidad Viento (km/h)')
        fig_viento.add_vline(x=ahora, line_width=2, line_dash="dash", line_color="white", annotation_text="AHORA")
        fig_viento.update_traces(line_color="#00CED1")
        st.plotly_chart(fig_viento, use_container_width=True)

# ------------------------------------------
# TAB 6: CENTRO DE ALERTAS Y SAT (CON GRÁFICOS PASADOS Y FUTUROS)
# ------------------------------------------
with tab6:
    st.subheader("Sala de Control Central")
    
    # --- SECCIÓN 1: BITÁCORA PASADA ---
    st.markdown("### 📜 Bitácora de Infracciones (Últimas 24 Horas)")
    st.write(f"Registro legal automatizado de contingencias normativas para **{contaminante_elegido}**.")
    
    alertas_pasadas = []
    hace_24h = ahora - pd.Timedelta(hours=24)
    
    for region, comunas in datos_totales.items():
        for comuna, sectores in comunas.items():
            for sector, serie in sectores.items():
                if not serie.empty:
                    serie_24h = serie[(serie.index >= hace_24h) & (serie.index <= ahora)]
                    if not serie_24h.empty:
                        horas_infraccion = (serie_24h > limite_actual).sum()
                        if horas_infraccion > 0:
                            peak = serie_24h.max()
                            ultimo_momento = serie_24h[serie_24h > limite_actual].index[-1]
                            alertas_pasadas.append({
                                "Fecha/Hora Último Peak": ultimo_momento.strftime("%Y-%m-%d %H:00"),
                                "Región": region,
                                "Comuna": comuna,
                                "Estación": sector,
                                "Peak (µg/m³)": round(peak, 1),
                                "Horas de Infracción": horas_infraccion
                            })
    
    if alertas_pasadas:
        df_alertas = pd.DataFrame(alertas_pasadas).sort_values("Fecha/Hora Último Peak", ascending=False).reset_index(drop=True)
        st.error(f"⚠️ Se detectaron vulneraciones a la normativa en **{len(alertas_pasadas)}** estaciones durante las últimas 24 horas.")
        st.dataframe(df_alertas, use_container_width=True, hide_index=True)
        
        # --- GRÁFICOS AUTOMÁTICOS PARA EL PASADO ---
        st.markdown("#### 📉 Evidencia Gráfica de Infracciones (Últimas 24h)")
        for alerta in alertas_pasadas:
            reg_graf = alerta["Región"]
            com_graf = alerta["Comuna"]
            est_graf = alerta["Estación"]
            
            # Recuperamos la serie de datos
            serie_grafico = datos_totales[reg_graf][com_graf][est_graf]
            df_g = serie_grafico.reset_index()
            df_g.columns = ['Fecha y Hora', 'Concentración']
            
            # Dibujamos el gráfico
            fig_pasado = px.line(
                df_g, x='Fecha y Hora', y='Concentración', 
                title=f"Infracción Registrada: {est_graf} ({com_graf})",
                color_discrete_sequence=["#FF7E00"] # Naranja para el pasado
            )
            fig_pasado.add_hline(y=limite_actual, line_dash="dot", line_color="red", annotation_text="Límite Legal")
            fig_pasado.add_vline(x=ahora, line_width=2, line_dash="dash", line_color="white", annotation_text="AHORA")
            
            # Acotamos el zoom para ver 48h hacia atrás y un poquito hacia adelante
            rango_inicio = ahora - pd.Timedelta(hours=48)
            rango_fin = ahora + pd.Timedelta(hours=12)
            fig_pasado.update_xaxes(range=[rango_inicio, rango_fin])
            
            st.plotly_chart(fig_pasado, use_container_width=True)

    else:
        st.success(f"✅ No se han registrado superaciones a la norma de {contaminante_elegido} durante las últimas 24 horas.")

    st.divider()

    # --- SECCIÓN 2: SISTEMA DE ALERTA TEMPRANA ---
    st.markdown("### 🔮 Sistema de Alerta Temprana (Próximas 72 Horas)")
    st.write("Escáner predictivo: Detecta zonas que superarán el límite legal en los próximos 3 días basándose en la dispersión meteorológica proyectada.")
    
    alertas_futuras = []
    
    for region, comunas in datos_totales.items():
        for comuna, sectores in comunas.items():
            for sector, serie in sectores.items():
                if not serie.empty:
                    serie_futura = serie[serie.index > ahora]
                    if not serie_futura.empty:
                        horas_peligro = (serie_futura > limite_actual).sum()
                        if horas_peligro > 0:
                            peak_futuro = serie_futura.max()
                            momento_peak = serie_futura[serie_futura == peak_futuro].index[0]
                            inicio_episodio = serie_futura[serie_futura > limite_actual].index[0]
                            
                            alertas_futuras.append({
                                "Inicio Proyectado": inicio_episodio.strftime("%Y-%m-%d %H:00"),
                                "Peak Máximo Estimado": momento_peak.strftime("%Y-%m-%d %H:00"),
                                "Región": region,
                                "Comuna": comuna,
                                "Estación": sector,
                                "Peak (µg/m³)": round(peak_futuro, 1),
                                "Horas en Riesgo": horas_peligro
                            })
                            
    if alertas_futuras:
        df_futuro = pd.DataFrame(alertas_futuras).sort_values("Inicio Proyectado").reset_index(drop=True)
        st.warning(f"⚠️ ¡ATENCIÓN! Se proyectan superaciones a la norma en **{len(alertas_futuras)}** estaciones para los próximos días.")
        st.dataframe(df_futuro, use_container_width=True, hide_index=True)
        
        # --- GRÁFICOS AUTOMÁTICOS PARA EL FUTURO ---
        st.markdown("#### 📈 Proyección Gráfica de Zonas en Riesgo")
        
        for alerta in alertas_futuras:
            reg_graf = alerta["Región"]
            com_graf = alerta["Comuna"]
            est_graf = alerta["Estación"]
            
            serie_grafico = datos_totales[reg_graf][com_graf][est_graf]
            df_g = serie_grafico.reset_index()
            df_g.columns = ['Fecha y Hora', 'Concentración']
            
            fig_sat = px.line(
                df_g, x='Fecha y Hora', y='Concentración', 
                title=f"Proyección Crítica: {est_graf} ({com_graf})",
                color_discrete_sequence=["#FF4B4B"] # Rojo claro para el futuro
            )
            fig_sat.add_hline(y=limite_actual, line_dash="dot", line_color="red", annotation_text="Límite Legal")
            fig_sat.add_vline(x=ahora, line_width=2, line_dash="dash", line_color="white", annotation_text="AHORA")
            
            # Zoom desde ayer hasta los 3 días proyectados
            rango_inicio = ahora - pd.Timedelta(hours=24)
            rango_fin = ahora + pd.Timedelta(hours=72)
            fig_sat.update_xaxes(range=[rango_inicio, rango_fin])
            
            st.plotly_chart(fig_sat, use_container_width=True)

    else:
        st.info("✅ El modelo predictivo indica que no habrá superaciones normativas en los próximos 3 días.")